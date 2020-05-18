import os
import glob
from pathlib import Path
import openpyxl
import csv

file_path = rf"C:\Users\P-67\Documents\csv_test\*.csv"
output_file_name = "output.xlsx"
csvfiles = glob.glob(file_path, recursive=False)
wb = openpyxl.Workbook()
for file in csvfiles:
    wb.create_sheet(os.path.splitext(os.path.basename(file))[0])
    wb.active = wb.sheetnames.index(os.path.splitext(os.path.basename(file))[0])
    ws = wb.active

    with open(file, encoding="shift-jis") as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
wb.save(output_file_name)